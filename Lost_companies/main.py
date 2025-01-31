"""
Все права защищены (c) 2023.
Данный скрипт анализирует «сделки» и «компании» в Excel, чтобы найти:
- сделки с редкими (потерянными) компаниями;
- компании без сделок;
- компании без документов.
Код не предназначен для свободного использования.
"""

import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from ttkthemes import ThemedTk
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def select_file1():
    """
    Диалог выбора основного файла (сделки).
    """
    file1_path.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx")]))


def select_file2():
    """
    Диалог выбора второго файла (компании).
    """
    file2_path.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx")]))


def clear_files():
    """
    Очищает пути к выбранным файлам и сбрасывает список «ответственных».
    """
    file1_path.set('')
    file2_path.set('')
    for widget in responsibles_frame.winfo_children():
        widget.destroy()
    responsible_vars.clear()


def is_valid_excel(file_path):
    """
    Проверяет, можно ли считать файл Excel.
    Возвращает True, если файл корректный.
    """
    try:
        pd.read_excel(file_path, engine='openpyxl')
        return True
    except Exception:
        return False


def process_files():
    """
    Анализирует второй файл (компании), находит всех «ответственных»
    и выводит список с чекбоксами.
    """
    if not file1_path.get() or not file2_path.get():
        messagebox.showerror("Ошибка", "Пожалуйста, выберите оба файла")
        return

    if not is_valid_excel(file1_path.get()):
        messagebox.showerror("Ошибка", "Файл1 (Сделки) не является корректным Excel-файлом")
        return

    if not is_valid_excel(file2_path.get()):
        messagebox.showerror("Ошибка", "Файл2 (Компании) не является корректным Excel-файлом")
        return

    file2 = pd.read_excel(file2_path.get(), engine='openpyxl')

    unique_responsibles = file2['Ответственный'].unique()
    filtered_responsibles = [resp for resp in unique_responsibles if resp.startswith('ОП') or resp.startswith('РО')]

    for widget in responsibles_frame.winfo_children():
        widget.destroy()

    responsible_vars.clear()
    for responsible in filtered_responsibles:
        var = tk.BooleanVar()
        ttk.Checkbutton(responsibles_frame, text=responsible, variable=var).pack(anchor=tk.W)
        responsible_vars[responsible] = var


def create_folder():
    """
    Создает папку на рабочем столе с названием «Потеряшки_YYYY-MM-DD», если она не существует.
    Возвращает путь к созданной/существующей папке.
    """
    desktop_path = os.path.join(os.environ['USERPROFILE'], 'Desktop')
    date_str = datetime.now().strftime("%Y-%m-%d")
    folder_name = os.path.join(desktop_path, f"Потеряшки_{date_str}")
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
    return folder_name


def filter_deals(base_folder, file1, selected_responsibles):
    """
    Для каждого ответственного из selected_responsibles:
    1) Фильтрует сделки по столбцу «Ответственный».
    2) Выделяет редкие компании (с количеством сделок < 4).
    3) Сохраняет результат в XLSX и визуально помечает редкие компании (#ead1dc).
    """
    for responsible in selected_responsibles:
        responsible_folder = os.path.join(base_folder, responsible)
        if not os.path.exists(responsible_folder):
            os.makedirs(responsible_folder)

        filtered_df = file1[file1['Ответственный'] == responsible][
            ['Ответственный', 'Название сделки', 'Компания']
        ].copy()

        company_counts = filtered_df['Компания'].value_counts()
        rare_companies = company_counts[company_counts < 4].index  # например, < 4 сделок

        # Помечаем редкие компании спецтегом
        filtered_df.loc[filtered_df['Компания'].isin(rare_companies), 'Компания'] = filtered_df['Компания'].apply(
            lambda x: f"{x} #ead1dc"
        )

        output_path = os.path.join(responsible_folder, f"Сделки_{responsible}.xlsx")
        filtered_df.to_excel(output_path, index=False)

        highlight_rare_companies(output_path)


def highlight_rare_companies(file_path):
    """
    Открывает сгенерированный XLSX и заменяет « #ead1dc» на цвет заливки ячейки.
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active
    fill = PatternFill(start_color='ead1dc', end_color='ead1dc', fill_type='solid')

    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if '#ead1dc' in str(cell.value):
                cell.value = cell.value.replace(' #ead1dc', '')
                cell.fill = fill

    workbook.save(file_path)


def filter_companies_no_deals(base_folder, file1, file2, selected_responsibles):
    """
    Находит компании с «Наличие док-в = Да», но у которых нет сделок.
    """
    for responsible in selected_responsibles:
        responsible_folder = os.path.join(base_folder, responsible)
        if not os.path.exists(responsible_folder):
            os.makedirs(responsible_folder)

        responsible_df = file2[
            (file2['Ответственный'] == responsible) & (file2['Наличие док-в'] == 'Да')
        ][['Ответственный', 'Название компании']]

        deals_df = file1[file1['Ответственный'] == responsible][['Ответственный', 'Компания']]
        no_deals_df = responsible_df[~responsible_df['Название компании'].isin(deals_df['Компания'])]

        output_path = os.path.join(responsible_folder, f"Компании без сделок_{responsible}.xlsx")
        no_deals_df.to_excel(output_path, index=False)


def filter_companies_no_docs(base_folder, file1, file2, selected_responsibles):
    """
    Находит компании, у которых «Наличие док-в = Нет», но при этом есть сделки.
    """
    for responsible in selected_responsibles:
        responsible_folder = os.path.join(base_folder, responsible)
        if not os.path.exists(responsible_folder):
            os.makedirs(responsible_folder)

        responsible_df = file2[
            (file2['Ответственный'] == responsible) & (file2['Наличие док-в'] == 'Нет')
        ][['Ответственный', 'Название компании']]

        deals_df = file1[file1['Ответственный'] == responsible][['Ответственный', 'Компания']]
        no_docs_df = responsible_df[responsible_df['Название компании'].isin(deals_df['Компания'])]

        output_path = os.path.join(responsible_folder, f"Компании без доков_{responsible}.xlsx")
        no_docs_df.to_excel(output_path, index=False)


def process_all():
    """
    Запускает всю цепочку: сделки с редкими компаниями, компании без сделок, компании без документов.
    """
    if not file1_path.get() or not file2_path.get():
        messagebox.showerror("Ошибка", "Пожалуйста, выберите оба файла")
        return

    file1 = pd.read_excel(file1_path.get(), engine='openpyxl')
    file2 = pd.read_excel(file2_path.get(), engine='openpyxl')

    selected_responsibles = [resp for resp, var in responsible_vars.items() if var.get()]
    if not selected_responsibles:
        messagebox.showerror("Ошибка", "Пожалуйста, выберите хотя бы одного ответственного")
        return

    base_folder = create_folder()

    filter_deals(base_folder, file1, selected_responsibles)
    filter_companies_no_deals(base_folder, file1, file2, selected_responsibles)
    filter_companies_no_docs(base_folder, file1, file2, selected_responsibles)


# ---------------------- Инициализация интерфейса ---------------------- #
root = ThemedTk(theme="arc")
root.title("Обработка сделок")

file1_path = tk.StringVar()
file2_path = tk.StringVar()
responsible_vars = {}

# Первый файл (Сделки)
ttk.Label(root, text="Файл1 - Сделки:").grid(row=0, column=0, padx=5, pady=5)
ttk.Entry(root, textvariable=file1_path, width=50).grid(row=0, column=1, padx=5, pady=5)
ttk.Button(root, text="Выбрать", command=select_file1).grid(row=0, column=2, padx=5, pady=5)

# Второй файл (Компании)
ttk.Label(root, text="Файл2 - Компании:").grid(row=1, column=0, padx=5, pady=5)
ttk.Entry(root, textvariable=file2_path, width=50).grid(row=1, column=1, padx=5, pady=5)
ttk.Button(root, text="Выбрать", command=select_file2).grid(row=1, column=2, padx=5, pady=5)

# Кнопки «Изучить» и «Очистить»
ttk.Button(root, text="Изучить", command=process_files).grid(row=2, column=1, padx=5, pady=5)
ttk.Button(root, text="Очистить", command=clear_files).grid(row=2, column=2, padx=5, pady=5)

# Фрейм для списка ответственных
responsibles_frame = ttk.Frame(root)
responsibles_frame.grid(row=3, column=0, columnspan=3, padx=5, pady=5)

# Отдельные кнопки по функционалу
ttk.Button(
    root, text="Сделки",
    command=lambda: filter_deals(
        create_folder(),
        pd.read_excel(file1_path.get(), engine='openpyxl'),
        [resp for resp, var in responsible_vars.items() if var.get()]
    )
).grid(row=4, column=1, padx=5, pady=5)

ttk.Button(
    root, text="Компании без сделок",
    command=lambda: filter_companies_no_deals(
        create_folder(),
        pd.read_excel(file1_path.get(), engine='openpyxl'),
        pd.read_excel(file2_path.get(), engine='openpyxl'),
        [resp for resp, var in responsible_vars.items() if var.get()]
    )
).grid(row=5, column=1, padx=5, pady=5)

ttk.Button(
    root, text="Компании без доков",
    command=lambda: filter_companies_no_docs(
        create_folder(),
        pd.read_excel(file1_path.get(), engine='openpyxl'),
        pd.read_excel(file2_path.get(), engine='openpyxl'),
        [resp for resp, var in responsible_vars.items() if var.get()]
    )
).grid(row=6, column=1, padx=5, pady=5)

ttk.Button(root, text="Выполнить все", command=process_all).grid(row=7, column=1, padx=5, pady=5)

root.mainloop()
