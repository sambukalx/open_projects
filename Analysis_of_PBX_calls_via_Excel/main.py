"""
Все права защищены (c) 2024. 
Данный скрипт предназначен для сбора и анализа данных о звонках из АТС.
Автор кода не предоставляет прав на использование или распространение данного ПО.
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import time, datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Side, Font, Border, Alignment, PatternFill
import pandas as pd
import os
import ttkbootstrap as ttkb
from tkinterdnd2 import DND_FILES, TkinterDnD
import time as tm
import tkinterdnd2.TkinterDnD
import tkinterdnd2.tkdnd


class AppData:
    """
    Хранилище глобальных данных приложения.
    """

    def __init__(self):
        self.wb = None
        self.successful_call_threshold_entry = ''
        self.original_file_path = ''
        self.processed_file_path = ''
        self.entry_widgets = {}
        self.successful_call_threshold = 50
        self.excluded_employees = {
            "XXX",
            "YYY",
            "ZZZ"
        }


data = AppData()

# Значения по умолчанию для критериев "Хорошо"/"Средне"
good = 27
norm = 25


# --------------------- Вспомогательные функции --------------------- #
def load_workbook_safe(filepath):
    """
    Безопасная загрузка книги Excel с обработкой ошибок.
    """
    try:
        return load_workbook(filename=filepath)
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось загрузить файл: {e}")
        return None


def get_border(style='thin'):
    """
    Возвращает объект границы ячейки с заданным стилем.
    """
    return Border(bottom=Side(style=style))


def get_fill(color):
    """
    Возвращает объект заливки ячейки заданным цветом.
    """
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def get_alignment(wrap_text=True):
    """
    Возвращает объект выравнивания ячейки (по умолчанию — перенос слов).
    """
    return Alignment(wrap_text=wrap_text)


def validate_and_parse_datetime(date_str, time_str, date_format='%Y-%m-%d', time_format='%H:%M:%S'):
    """
    Безопасно парсит дату и время в один объект datetime.
    Можно настроить форматы при необходимости.
    """
    parsed_date = (date_str.date() if isinstance(date_str, datetime)
                   else datetime.strptime(date_str, date_format).date())
    parsed_time = (time_str if isinstance(time_str, time)
                   else datetime.strptime(time_str, time_format).time())
    return datetime.combine(parsed_date, parsed_time)


def convert_time_to_minutes(t):
    """
    Преобразует объект time в минуты (float). Если t пустой, возвращает 0.
    """
    if isinstance(t, time):
        return t.hour * 60 + t.minute + t.second / 60.0
    return float(t) if t else 0


def add_bottom_border(sheet, start_row, end_row):
    """
    Добавляет нижнюю границу для выделенного диапазона строк на листе.
    """
    border_style = get_border()
    for row in range(start_row, end_row + 1):
        for cell in sheet[row]:
            if cell.value is not None or (
                    cell.fill and cell.fill.start_color and cell.fill.start_color.index != '00000000'
            ):
                cell.border = border_style


# --------------------- Обработчики интерфейса --------------------- #
def open_file():
    """
    Открывает диалог выбора файла и загружает Excel-файл.
    """
    data.original_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if data.original_file_path:
        data.wb = load_workbook_safe(data.original_file_path)
        if data.wb:
            status_label.config(text="Файл успешно загружен")
        else:
            status_label.config(text="Ошибка загрузки файла")


def study_file():
    """
    Изучает выбранный файл и собирает имена сотрудников. 
    Строит GUI-форму для ввода порогов звонков (хорошо/средне/плохо).
    """
    if data.wb and data.original_file_path:
        unique_names = set()
        sheet = data.wb["Sheet1"]

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row,
                                   min_col=3, max_col=3, values_only=True):
            name = row[0]
            if name and name.strip() not in data.excluded_employees:
                unique_names.add(name)

        if unique_names:
            # Очищаем предыдущие виджеты в скроллируемой области
            for widget in frame.winfo_children():
                widget.destroy()

            data.entry_widgets = {}

            # Шапка таблицы
            tk.Label(frame, text="Сотрудник", fg="black").grid(
                row=0, column=0, padx=(10, 0), pady=(5, 0), sticky="w"
            )
            tk.Label(frame, text="Хорошо", fg="green").grid(
                row=0, column=1, padx=(10, 0), pady=(5, 0), sticky="w"
            )
            tk.Label(frame, text="Средне", fg="orange").grid(
                row=0, column=2, padx=(10, 0), pady=(5, 0), sticky="w"
            )
            tk.Label(frame, text="Плохо", fg="red").grid(
                row=0, column=3, padx=(10, 0), pady=(5, 0), sticky="w"
            )

            # Для каждого сотрудника создаём ряд с полями ввода
            for i, name in enumerate(sorted(unique_names), start=1):
                label = tk.Label(frame, text=name)
                label.grid(row=i, column=0, sticky="w", padx=(10, 0), pady=(5, 0))

                entries = []
                for j in range(3):
                    entry = tk.Entry(
                        frame,
                        validate="key",
                        validatecommand=(root.register(validate_input), "%P")
                    )
                    entry.grid(row=i, column=j + 1, padx=(1, 0), pady=(5, 0), sticky="ew")
                    entries.append(entry)

                data.entry_widgets[name] = entries

                # По умолчанию вставляем значения
                data.entry_widgets[name][0].insert(0, good)  # Хорошо
                data.entry_widgets[name][1].insert(0, norm)  # Средне
                data.entry_widgets[name][2].insert(0, norm - 1)  # Плохо

            # Добавляем поле порога успешного звонка
            threshold_row = len(unique_names) + 1
            tk.Label(frame, text="Порог успешного звонка (сек.):").grid(
                row=threshold_row, column=0, padx=(10, 0), pady=(10, 0), sticky="w"
            )
            data.successful_call_threshold_entry = tk.Entry(frame)
            data.successful_call_threshold_entry.insert(0, str(data.successful_call_threshold))
            data.successful_call_threshold_entry.grid(
                row=threshold_row, column=1, padx=(1, 0), pady=(10, 0), sticky="w"
            )

            # Обновляем область скролла
            canvas.update_idletasks()
            canvas.config(scrollregion=canvas.bbox("all"))

            status_label.config(text="Уникальные имена отображены")
            root.after(300, lambda: status_label.config(text=""))
        else:
            status_label.config(text="В столбце C нет имен")
    else:
        status_label.config(text="Сначала выберите файл для анализа")


# ------------------- Основные функции анализа ------------------- #
def create_analysis_sheet(wb_new, sheet, new_sheet, min_successful_duration):
    """
    Создаёт сводную таблицу по среднему времени ожидания/успешных звонков.
    """
    employees = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[2]
        if not name or name in data.excluded_employees:
            continue

        waiting_time = convert_time_to_minutes(row[9])  # время ожидания
        call_duration = convert_time_to_minutes(row[10])  # продолжительность

        if name not in employees:
            employees[name] = {'waiting_times': [], 'successful_durations': []}

        employees[name]['waiting_times'].append(waiting_time)

        if call_duration >= min_successful_duration:
            employees[name]['successful_durations'].append(call_duration)

    total_successful_duration = 0
    total_successful_count = 0
    row_num = 2

    for name, data_item in sorted(employees.items()):
        avg_waiting = sum(data_item['waiting_times']) / len(data_item['waiting_times'])
        if data_item['successful_durations']:
            avg_successful_call = sum(data_item['successful_durations']) / len(data_item['successful_durations'])
        else:
            avg_successful_call = 0

        total_successful_duration += sum(data_item['successful_durations'])
        total_successful_count += len(data_item['successful_durations'])

        new_sheet.append([
            name,
            f"{int(avg_waiting)}:{int((avg_waiting % 1) * 60):02d}",
            f"{int(avg_successful_call)}:{int((avg_successful_call % 1) * 60):02d}"
        ])
        row_num += 1

    add_bottom_border(new_sheet, 2, row_num - 1)

    if total_successful_count:
        overall_avg_successful = total_successful_duration / total_successful_count
        new_sheet['C1'] = f"{int(overall_avg_successful)}:{int((overall_avg_successful % 1) * 60):02d}"
    else:
        new_sheet['C1'] = "0:00"


def create_call_summary_sheet(wb_new, sheet, new_sheet, min_successful_duration):
    """
    Создаёт таблицу подсчёта типов звонков (входящие, исходящие, пропущенные и т.д.).
    """
    call_type_column = 'A'
    name_column = 'C'
    duration_column = 'K'
    employees_calls = {}

    for row in sheet.iter_rows(min_row=10, values_only=True):
        call_type = row[ord(call_type_column) - 65]
        name = row[ord(name_column) - 65]
        duration = row[ord(duration_column) - 65]

        if not name or not call_type or name in data.excluded_employees:
            continue

        if name not in employees_calls:
            employees_calls[name] = {
                'входящий': 0,
                'исходящий': 0,
                'пропущенный': 0,
                'неуспешный исходящий': 0,
                'успешный': 0
            }

        if call_type in employees_calls[name]:
            employees_calls[name][call_type] += 1

        duration_minutes = convert_time_to_minutes(duration)
        if duration_minutes >= min_successful_duration:
            employees_calls[name]['успешный'] += 1

    row_num = 16
    for name, calls in sorted(employees_calls.items()):
        if name in data.excluded_employees:
            continue
        new_sheet[f'A{row_num}'] = name
        new_sheet[f'B{row_num}'] = calls['входящий']
        new_sheet[f'C{row_num}'] = calls['исходящий']
        new_sheet[f'D{row_num}'] = calls['пропущенный']
        new_sheet[f'E{row_num}'] = calls['неуспешный исходящий']
        new_sheet[f'F{row_num}'] = sum(calls.values()) - calls['успешный']
        new_sheet[f'G{row_num}'] = calls['успешный']
        row_num += 1

    add_bottom_border(new_sheet, 16, row_num - 1)


def create_daily_calls_report(wb_new, sheet, new_sheet, start_date, end_date, employees, min_successful_duration):
    """
    Формирует помесячный (или подневной) отчёт с разбивкой по датам, считая
    входящие/исходящие/пропущенные/успешные звонки для каждого сотрудника.
    """
    date_range = pd.date_range(start=start_date, end=end_date)

    calls_by_date = {
        name: {date.strftime("%m/%d/%Y"): {'calls': [], 'successful_calls': 0} for date in date_range}
        for name in employees if name not in data.excluded_employees
    }
    call_types = ['входящий', 'исходящий', 'пропущенный', 'неуспешный исходящий']

    for row in sheet.iter_rows(min_row=10, values_only=True):
        call_date = row[7]
        if isinstance(call_date, datetime) and start_date <= call_date <= end_date:
            name = row[2]
            call_type = row[0]
            duration = convert_time_to_minutes(row[10])
            call_date_str = call_date.strftime("%m/%d/%Y")

            if name and name in calls_by_date:
                calls_by_date[name][call_date_str]['calls'].append(call_type)
                # Если звонок длится >= min_successful_duration (минут), считаем его успешным
                if duration >= min_successful_duration:
                    calls_by_date[name][call_date_str]['successful_calls'] += 1

    row_num = 38
    side_border = get_border()

    for date in date_range:
        # Печатаем дату
        date_cell = new_sheet.cell(row=row_num, column=1, value=date.strftime("%m/%d/%Y"))
        date_cell.border = side_border
        row_num += 1

        for name, dates in calls_by_date.items():
            if name in data.excluded_employees:
                continue

            name_cell = new_sheet.cell(row=row_num, column=1, value=name)
            name_cell.border = side_border

            # Подсчитываем типы звонков по конкретной дате
            call_counts = {ct: dates[date.strftime("%m/%d/%Y")]['calls'].count(ct) for ct in call_types}
            successful_calls = dates[date.strftime("%m/%d/%Y")]['successful_calls']

            # Заполняем ячейки с количеством каждого типа звонка
            for i, ct in enumerate(call_types, start=2):
                cell = new_sheet.cell(row=row_num, column=i, value=call_counts[ct])
                cell.border = side_border

            # Общее количество всех звонков
            total_calls_cell = new_sheet.cell(row=row_num, column=len(call_types) + 2,
                                              value=sum(call_counts.values()))
            total_calls_cell.border = side_border

            # Успешные звонки
            successful_calls_cell = new_sheet.cell(row=row_num, column=len(call_types) + 3,
                                                   value=successful_calls)
            successful_calls_cell.border = side_border

            # Окраска ячейки по порогам (хорошо/средне/плохо)
            green_threshold, yellow_threshold, red_threshold = [
                int(entry.get()) for entry in data.entry_widgets[name]
            ]

            if successful_calls >= green_threshold:
                color = "00FF00"  # зелёный
            elif green_threshold > successful_calls >= yellow_threshold:
                color = "FFFF00"  # жёлтый
            else:
                color = "FF0000"  # красный

            target_cell = new_sheet.cell(row=row_num, column=8)
            target_cell.fill = get_fill(color)

            row_num += 1

        row_num += 1

    add_bottom_border(new_sheet, 28, row_num - 1)

    return calls_by_date, row_num


def create_break_analysis(wb_new, sheet, new_sheet, start_date, end_date, employees, date_range, calls_by_date,
                          row_num):
    """
    Анализ длительных перерывов (больше 1ч 10мин, но меньше 8ч).
    """
    calls_by_employee = {}
    added_entries = set()  # Чтобы исключить дубли 

    # Собираем все даты/время звонков для каждого сотрудника
    for row in sheet.iter_rows(min_row=2, values_only=True):
        employee_name = row[2]
        call_date_str = row[7]
        call_time_str = row[8]

        if employee_name and employee_name not in data.excluded_employees and call_date_str and call_time_str:
            call_datetime = validate_and_parse_datetime(call_date_str, call_time_str)
            if call_datetime:
                if employee_name not in calls_by_employee:
                    calls_by_employee[employee_name] = []
                calls_by_employee[employee_name].append(call_datetime)

    start_row = 2
    start_column = 7

    for employee in employees:
        if employee not in data.excluded_employees:
            large_breaks_count_by_date = {date.strftime("%m/%d/%Y"): 0 for date in date_range}
            times = calls_by_employee.get(employee, [])
            times.sort()

            longest_break = None
            start_of_longest = None
            end_of_longest = None
            large_breaks = 0

            for i in range(1, len(times)):
                current_break = times[i] - times[i - 1]
                # Считаем «длинным перерывом» от 1 ч 10 мин до 8 часов
                if timedelta(hours=1, minutes=10) < current_break < timedelta(hours=8):
                    large_breaks += 1
                    break_date = times[i - 1].date().strftime("%m/%d/%Y")

                    if break_date in large_breaks_count_by_date:
                        large_breaks_count_by_date[break_date] += 1

                    # Запоминаем самый длинный перерыв
                    if not longest_break or current_break > longest_break:
                        longest_break = current_break
                        start_of_longest = times[i - 1]
                        end_of_longest = times[i]

            if employee in calls_by_date:
                for date_str in large_breaks_count_by_date:
                    if date_str in calls_by_date[employee]:
                        calls_by_date[employee][date_str]['large_breaks'] = large_breaks_count_by_date[date_str]

            # Заполняем таблицу на листе "Анализ"
            if longest_break:
                entry = (employee, start_of_longest, end_of_longest, longest_break)
                if entry not in added_entries:
                    added_entries.add(entry)

                    sheet1 = new_sheet
                    sheet1.cell(row=start_row, column=start_column, value=employee)
                    sheet1.cell(row=start_row, column=start_column + 2, value=start_of_longest)
                    sheet1.cell(row=start_row, column=start_column + 3, value=end_of_longest)
                    sheet1.cell(row=start_row, column=start_column + 4, value=longest_break)

                    # Стили заливки
                    yellow_fill = get_fill("FFFF00")
                    red_fill = get_fill("FF0000")
                    green_fill = get_fill("00FF00")

                    # Пример простого условия цвета (можно доработать под свою логику)
                    fill_color = (
                        yellow_fill if len(times) == 1
                        else red_fill if len(times) > 1
                        else green_fill
                    )
                    sheet1.cell(row=start_row, column=8).fill = fill_color

                    if large_breaks:
                        new_sheet.cell(row=start_row, column=12, value=large_breaks)

                    start_row += 1

    add_bottom_border(new_sheet, 2, row_num - 1)


def apply_formatting_to_sheet(new_sheet, start_date_str, end_date_str):
    """
    Настраивает внешний вид итогового листа (ширину столбцов, шрифты, закреплённые области).
    """
    columns_to_update = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'I', 'J']
    for column in columns_to_update:
        current_width = new_sheet.column_dimensions[column].width
        new_sheet.column_dimensions[column].width = current_width * 2 if current_width else 20

    bold_font = Font(bold=True)
    thin_border = get_border()

    for cell in [
        'A1', 'B1', 'C1',
        'A15', 'B15', 'C15', 'D15', 'E15', 'F15', 'G15',
        'E2', 'E3', 'E4', 'G1', 'I1', 'J1', 'K1', 'L1'
    ]:
        new_sheet[cell].font = bold_font
        new_sheet[cell].border = thin_border

    new_sheet['A1'].alignment = get_alignment()
    new_sheet['B1'].alignment = get_alignment()
    new_sheet['C1'].alignment = get_alignment()

    new_sheet.freeze_panes = 'H16'
    new_sheet.sheet_view.zoomScale = 85

    # Пытаемся распарсить даты для красивого описания
    try:
        start_date = datetime.strptime(start_date_str, "%m/%d/%Y")
    except ValueError:
        start_date = None

    try:
        end_date = datetime.strptime(end_date_str, "%m/%d/%Y")
    except ValueError:
        end_date = None

    new_sheet['E2'] = f'Начало {start_date_str}'
    new_sheet['E3'] = f'Конец {end_date_str}'
    if start_date and end_date:
        days_count = (end_date - start_date).days + 1
        new_sheet['E4'] = f'Данные за {days_count} дней'

    new_sheet['G1'] = 'Сотрудник'
    new_sheet['I1'] = 'Начало перерыва'
    new_sheet['J1'] = 'Конец перерыва'
    new_sheet['K1'] = 'Время'
    new_sheet['L1'] = 'Кол-во'

    new_sheet['A15'] = 'Сотрудник'
    new_sheet['B15'] = 'входящий'
    new_sheet['C15'] = 'исходящий'
    new_sheet['D15'] = 'неотвеченный'
    new_sheet['E15'] = 'неуспешный исходящий'
    new_sheet['F15'] = 'общее'
    new_sheet['G15'] = 'успешный'


def process_file():
    """
    Основной метод «Обработать»:
    1) Берёт threshold успешного звонка.
    2) Создаёт новый лист «Анализ» и заполняет его сводками.
    3) Сохраняет новую книгу с постфиксом "_processed.xlsx".
    """
    if data.wb and data.original_file_path:
        try:
            data.successful_call_threshold = int(data.successful_call_threshold_entry.get())
        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректное значение порога времени успешного звонка (число).")
            return

        min_successful_duration = data.successful_call_threshold / 60.0
        new_file_path = data.original_file_path.replace(".xlsx", "_processed.xlsx")
        wb_new = load_workbook_safe(data.original_file_path)
        if not wb_new:
            return

        if "Анализ" in wb_new.sheetnames:
            if not messagebox.askyesno(
                    "Предупреждение",
                    "Лист 'Анализ' уже существует. Хотите перезаписать его?"
            ):
                return
            wb_new.remove(wb_new["Анализ"])

        new_sheet = wb_new.create_sheet(title="Анализ")
        new_sheet.append(["Сотрудник", "Среднее время ожидания", "Среднее время разговора успешного звонка"])

        progress_bar["maximum"] = 100
        progress_bar["value"] = 0
        status_label.config(text="Обработка файла...")

        # Имитация работы через задержку (для показа прогресс-бара)
        for i in range(5):
            progress_bar["value"] += 20
            root.update_idletasks()
            tm.sleep(1)

        sheet = wb_new["Sheet1"]
        start_date_str = sheet["B6"].value
        end_date_str = sheet["B7"].value

        # Пробуем разные форматы для start_date
        def try_parse_date(date_str, patterns):
            for pattern in patterns:
                try:
                    return datetime.strptime(date_str, pattern)
                except ValueError:
                    pass
            return None

        start_date = try_parse_date(start_date_str, ["%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"])
        end_date = try_parse_date(end_date_str, ["%Y-%d-%m", "%Y/%d/%m", "%d/%m/%Y", "%m/%d/%Y"])

        # Проверка корректности дат
        if not start_date or not end_date:
            messagebox.showerror("Ошибка", "Неверный формат дат начала/окончания в ячейках B6/B7.")
            return
        if start_date > end_date:
            messagebox.showerror("Ошибка", "Дата начала не может быть позже даты окончания.")
            return

        # Составляем список сотрудников (без исключённых)
        employees = [
            row[2] for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[2] and row[2] not in data.excluded_employees
        ]

        # Формирование частей отчёта
        create_analysis_sheet(wb_new, sheet, new_sheet, min_successful_duration)
        create_call_summary_sheet(wb_new, sheet, new_sheet, min_successful_duration)
        calls_by_date, row_num = create_daily_calls_report(
            wb_new, sheet, new_sheet,
            start_date, end_date,
            employees,
            min_successful_duration
        )
        create_break_analysis(
            wb_new, sheet, new_sheet,
            start_date, end_date,
            employees,
            pd.date_range(start=start_date, end=end_date),
            calls_by_date,
            row_num
        )
        apply_formatting_to_sheet(new_sheet, start_date_str, end_date_str)

        # Сохранение обработанного файла
        try:
            wb_new.save(filename=new_file_path)
            status_label.config(text="Файл успешно обработан")
            data.processed_file_path = new_file_path
            progress_bar["value"] = 100
        except PermissionError as e:
            status_label.config(text=f"Ошибка сохранения файла: {e}")
        finally:
            # Разблокируем кнопки "Скачать" и "Открыть"
            download_button.config(state=tk.NORMAL)
            open_button.config(state=tk.NORMAL)
    else:
        status_label.config(text="Файл не загружен или не найден")


def on_drop(event):
    """
    Позволяет загрузить файл, просто перетащив его на окно приложения.
    """
    file_path = event.data.replace('{', '').replace('}', '')
    data.original_file_path = file_path
    if data.original_file_path:
        data.wb = load_workbook_safe(data.original_file_path)
        if data.wb:
            status_label.config(text="Файл успешно загружен")
        else:
            status_label.config(text="Ошибка загрузки файла")


def download_file():
    """
    Сохраняет обработанный файл под выбранным пользователем именем.
    """
    if data.processed_file_path:
        download_path = filedialog.asksaveasfilename(
            initialdir=os.getcwd(),
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=os.path.basename(data.processed_file_path)
        )
        if download_path:
            os.replace(data.processed_file_path, download_path)
            status_label.config(text="Файл успешно сохранен")


def open_processed_file():
    """
    Открывает обработанный файл, если он существует.
    """
    if data.processed_file_path:
        if os.path.exists(data.processed_file_path):
            os.startfile(data.processed_file_path)
        else:
            messagebox.showerror("Ошибка", "Файл не найден")
    else:
        messagebox.showerror("Ошибка", "Файл не был обработан")


def validate_input(value):
    """
    Валидатор для Entry (допускаем только цифры).
    """
    return value.isdigit() or value == ""


def on_frame_configure(event):
    """
    Обновляет видимую область canvas при изменении размеров вложенного фрейма.
    """
    canvas.configure(scrollregion=canvas.bbox("all"))


def create_main_window():
    """
    Создаёт главное окно приложения, настраивает вкладки, кнопки и скролл.
    """
    global root, canvas, frame, status_label, open_button, download_button, progress_bar

    # Окно с поддержкой Drag-and-Drop
    root = TkinterDnD.Tk()
    root.title("Приложение для работы с файлом Excel")
    root.style = ttkb.Style("solar")

    # Вкладки (Notebook)
    notebook = ttkb.Notebook(root, bootstyle="info")
    notebook.pack(fill="both", expand=True, padx=10, pady=10)

    # Вкладка «Файл»
    frame_file_operations = ttkb.Frame(notebook)
    notebook.add(frame_file_operations, text="Файл")

    # Панель инструментов
    toolbar = ttkb.Frame(frame_file_operations)
    toolbar.pack(side="top", fill="x", pady=5)

    open_button = ttkb.Button(toolbar, text="Выбрать файл .xlsx", command=open_file, bootstyle="primary-outline")
    open_button.pack(side="left", padx=5)

    study_button = ttkb.Button(toolbar, text="Изучить", command=study_file, bootstyle="success-outline")
    study_button.pack(side="left", padx=5)

    process_button = ttkb.Button(toolbar, text="Обработать", command=process_file, bootstyle="warning-outline")
    process_button.pack(side="left", padx=5)

    download_button = ttkb.Button(toolbar, text="Скачать", command=download_file, bootstyle="info-outline")
    download_button.pack(side="left", padx=5)
    download_button.config(state=tk.DISABLED)

    open_button = ttkb.Button(toolbar, text="Открыть", command=open_processed_file, bootstyle="danger-outline")
    open_button.pack(side="left", padx=5)
    open_button.config(state=tk.DISABLED)

    # Статус-лейбл и прогресс-бар
    status_label = ttkb.Label(frame_file_operations, text="")
    status_label.pack(pady=10)

    progress_bar = ttkb.Progressbar(frame_file_operations, orient="horizontal", mode="determinate", bootstyle="info")
    progress_bar.pack(fill="x", padx=10, pady=10)

    # Drag-and-Drop
    root.drop_target_register(DND_FILES)
    root.dnd_bind('<<Drop>>', on_drop)

    # Вкладка «Данные» (со скроллируемой областью)
    frame_data_view = ttkb.Frame(notebook)
    notebook.add(frame_data_view, text="Данные")

    canvas = tk.Canvas(frame_data_view, width=500, height=300)
    scrollbar = ttkb.Scrollbar(frame_data_view, orient="vertical", command=canvas.yview)
    frame = ttkb.Frame(canvas)

    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.create_window((0, 0), window=frame, anchor="nw")

    frame.bind("<Configure>", on_frame_configure)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")


# Точка входа
if __name__ == "__main__":
    create_main_window()
    root.mainloop()
