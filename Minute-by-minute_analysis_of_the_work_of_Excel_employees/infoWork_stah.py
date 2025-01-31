"""
Все права защищены (c) 2024. 
Данный скрипт предназначен для обработки JSON файла.
Автор кода не предоставляет прав на использование или распространение данного ПО.
"""

import os
import json
from openpyxl import load_workbook
from datetime import time
from colorama import init, Fore
from colorama import Style

init(autoreset=True)
info_work_stah = {}


# Функция для объединения всех сотрудников в один список
def get_all_employees(department_dict):
    """
    :param department_dict: Словарь, где ключами являются названия отделов, а значениями — списки имен сотрудников.
    :return: Список, содержащий всех сотрудников всех отделов вместе взятых.
    """
    print(Fore.GREEN + 'Идет объединение всех сотрудников в один список')
    all_employees = []
    for department in department_dict.values():
        all_employees.extend(department)
    print(Fore.BLUE + 'get_all_employees в infoWork_stah.py выполнено\n')
    return all_employees


# Функция для обработки данных из файла worktime
def process_worktime_xlsx(file_path, employee_name):
    """
    :param file_path: Путь к файлу Excel, содержащему данные о рабочем времени сотрудников.
    :param employee_name: имя сотрудника, данные которого обрабатываются.
    :return: Нет. Функция добавляет обработанные данные о рабочем времени во внешнюю структуру info_work_stah.
    """
    wb = load_workbook(file_path)
    sheet = wb.active
    filename = os.path.basename(file_path)
    year_month = filename.split('_')[0]
    year, month = year_month.split('-')
    employee_in_file = sheet['C2'].value
    if employee_in_file != employee_name:
        print(f"Несоответствие имени: в файле {employee_in_file}, в папке {employee_name}")
        return
    if employee_name not in info_work_stah:
        info_work_stah[employee_name] = {
            'pc_time': {},
            'enterexit': {},
            'time_act': {}
        }
    if month not in info_work_stah[employee_name]['pc_time']:
        info_work_stah[employee_name]['pc_time'][month] = {}
    for col in range(4, sheet.max_column + 1):
        day_cell = sheet.cell(row=1, column=col)
        time_cell = sheet.cell(row=2, column=col)
        if isinstance(day_cell.value, str) and day_cell.value.split()[0].isdigit():
            day = int(day_cell.value.split()[0])
            time_value = time_cell.value
            if isinstance(time_value, str):
                if 'ч' in time_value:
                    hours, minutes = time_value.split('ч')
                    mean = f"{hours.strip()} часов {minutes.strip()} минут"
                elif time_value in ['П', 'В']:
                    mean = time_value
                info_work_stah[employee_name]['pc_time'][month][day] = {'mean': mean}


# Функция для обработки enterexit файлов
def process_entrexit_xlsx(file_path, employee_name):
    """
    :param file_path: Путь к файлу Excel, содержащему данные о рабочем времени.
    :param employee_name: имя сотрудника, данные которого необходимо обработать.
    :return: None
    """
    wb = load_workbook(file_path)
    sheet = wb.active
    # Получаем месяц и год из названия файла
    filename = os.path.basename(file_path)
    year_month = filename.split('_')[0]
    year, month = year_month.split('-')
    employee_in_file = sheet['C2'].value
    if employee_in_file != employee_name:
        print(f"Несоответствие имени: в файле {employee_in_file}, в папке {employee_name}")
        return
    if employee_name not in info_work_stah:
        info_work_stah[employee_name] = {
            'pc_time': {},
            'enterexit': {},
            'time_act': {}
        }
    if month not in info_work_stah[employee_name]['enterexit']:
        info_work_stah[employee_name]['enterexit'][month] = {}
    for col in range(4, sheet.max_column, 2):
        start_cell = sheet.cell(row=1, column=col)
        end_cell = sheet.cell(row=1, column=col + 1)
        start_time_cell = sheet.cell(row=2, column=col)
        end_time_cell = sheet.cell(row=2, column=col + 1)
        if isinstance(start_cell.value, str) and 'Начало' in start_cell.value:
            day = int(start_cell.value.split()[1])
            start_time = start_time_cell.value
            if isinstance(start_time, time):
                start_time = start_time.strftime("%H:%M")
            elif start_time in ['П', 'В']:
                start_time = start_time
            end_time = end_time_cell.value
            if isinstance(end_time, time):
                end_time = end_time.strftime("%H:%M")
            elif end_time in ['П', 'В']:
                end_time = end_time
            info_work_stah[employee_name]['enterexit'][month][day] = {
                'start': {'means': start_time},
                'end': {'meane': end_time}
            }


# Функция для обработки acttime файлов
def process_acttime_xlsx(file_path, employee_name):
    """
    :param file_path: Путь к файлу Excel, содержащему время активности. Имя файла должно соответствовать шаблону «ГГГГ-ММ_employee.xlsx».
    :param employee_name: Имя сотрудника. Он должен совпадать с тем, который присутствует в ячейке «C2» файла Excel.
    :return: Нет. Он обновляет глобальную структуру «info_work_stah» временем активности для указанного сотрудника и месяца.
    """
    wb = load_workbook(file_path)
    sheet = wb.active
    filename = os.path.basename(file_path)
    year_month = filename.split('_')[0]
    year, month = year_month.split('-')
    employee_in_file = sheet['C2'].value
    if employee_in_file != employee_name:
        print(f"Несоответствие имени: в файле {employee_in_file}, в папке {employee_name}")
        return
    if employee_name not in info_work_stah:
        info_work_stah[employee_name] = {
            'pc_time': {},
            'enterexit': {},
            'time_act': {}
        }
    if month not in info_work_stah[employee_name]['time_act']:
        info_work_stah[employee_name]['time_act'][month] = {}
    for col in range(4, sheet.max_column + 1):
        day_cell = sheet.cell(row=1, column=col)
        act_time_cell = sheet.cell(row=2, column=col)
        if isinstance(day_cell.value, str) and day_cell.value.split()[0].isdigit():
            day = int(day_cell.value.split()[0])
            act_time_value = act_time_cell.value
            if isinstance(act_time_value, time):
                act_time_value = act_time_value.strftime("%H:%M")
            elif act_time_value in ['П', 'В']:
                act_time_value = act_time_value
            info_work_stah[employee_name]['time_act'][month][day] = {'meant': act_time_value}


# Функция для сканирования всех подкаталогов
def scan_folders(root_path, all_employees):
    """
    :param root_path: Путь к корневому каталогу, в котором расположены сканируемые папки.
    :param all_employees: список всех имен сотрудников, которые нужно искать в корневом каталоге.
    :return: None
    """
    print(Fore.GREEN + f"Сканирование всех папок в директории: {root_path}")
    for dirpath, dirnames, filenames in os.walk(root_path):
        for employee_name in all_employees:
            employee_folder = os.path.join(dirpath, employee_name)
            if os.path.exists(employee_folder):
                for file_name in os.listdir(employee_folder):
                    file_path = os.path.join(employee_folder, file_name)
                    # Обработка файлов worktime
                    if file_name.endswith('.xlsx') and 'worktime' in file_name:
                        process_worktime_xlsx(file_path, employee_name)
                    # Обработка файлов entrexit
                    elif file_name.endswith('.xlsx') and 'entrexit' in file_name:
                        process_entrexit_xlsx(file_path, employee_name)
                    # Обработка файлов acttime
                    elif file_name.endswith('.xlsx') and 'acttime' in file_name:
                        process_acttime_xlsx(file_path, employee_name)


# Красивый вывод info_work_stah
def print_pretty_info_work_stah(info_work_stah):
    """
    :param info_work_stah: Словарь, содержащий информацию о сотрудниках. Структура словаря должна быть следующей:
        {
            "employee_name1": {
                "pc_time": { "month": { "day": {"mean": value} }},
                "enterexit": { "month": { "day": { "start": {"means": value}, "end": {"meane": value} }}},
                "time_act": { "month": { "day": {"meant": value} }},
            },
            "employee_name2": {
                ...
            },
            ...
        }
    :return: None. Функция выводит на консоль отформатированную информацию о сотруднике.
    """
    print("Отчет по сотрудникам:\n")
    for employee_name, employee_data in info_work_stah.items():
        print(f"Сотрудник: {employee_name}")
        # Вывод времени работы ПК
        if 'pc_time' in employee_data and employee_data['pc_time']:
            print("  Время работы компьютера:")
            for month, days in employee_data['pc_time'].items():
                print(f"    Месяц: {month}")
                for day, day_data in days.items():
                    mean_time = day_data.get('mean', 'Нет данных')
                    print(f"      День {day}: {mean_time}")
        # Вывод данных по времени входа и выхода
        if 'enterexit' in employee_data and employee_data['enterexit']:
            print("  Время входа/выхода:")
            for month, days in employee_data['enterexit'].items():
                print(f"    Месяц: {month}")
                for day, day_data in days.items():
                    start_time = day_data.get('start', {}).get('means', 'Нет данных')
                    end_time = day_data.get('end', {}).get('meane', 'Нет данных')
                    print(f"      День {day}: Вход - {start_time}, Выход - {end_time}")
        # Вывод активного времени
        if 'time_act' in employee_data and employee_data['time_act']:
            print("  Активное время:")
            for month, days in employee_data['time_act'].items():
                print(f"    Месяц: {month}")
                for day, day_data in days.items():
                    active_time = day_data.get('meant', 'Нет данных')
                    print(f"      День {day}: Активное время - {active_time}")
        print("\n")


def print_pretty_info_work_stah1(info_work_stah):
    """
    :param info_work_stah: данные для печати в красивом формате JSON.
    :return: None
    """
    print("Результат info_work_stah (красивый формат):")
    print(json.dumps(info_work_stah, indent=4, ensure_ascii=False))

