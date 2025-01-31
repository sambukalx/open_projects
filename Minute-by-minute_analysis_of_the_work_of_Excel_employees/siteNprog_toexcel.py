"""
Все права защищены (c) 2024. 
Данный скрипт предназначен для работы с Excel файлом.
Автор кода не предоставляет прав на использование или распространение данного ПО.
"""

from colorama import init, Fore
from colorama import Style
from openpyxl import load_workbook
from lxml import etree
import pandas as pd


def load_program_data(output_file_path_prog):
    """
    Загрузка данных о программах из XML файла и преобразование их в DataFrame.

    :param output_file_path_prog: Путь к XML файлу с данными.
    :return: DataFrame с данными о программах.
    """
    print(Fore.GREEN + 'Идет загрузка данных и программах из xml фала')
    print(output_file_path_prog)
    """Загрузить данные о программах из XML файла."""
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(output_file_path_prog, parser)
    root = tree.getroot()
    program_data = []
    for user in root.iter('user'):
        fio = user.find('fio')
        fio = fio.text if fio is not None else 'Неизвестно'
        for item in user.findall('item'):
            desc = item.find('desc')
            desc = desc.text if desc is not None else 'Неизвестно'
            stime = item.find('stime')
            stime = stime.text if stime is not None else 'Неизвестно'
            if stime != 'Неизвестно':
                parts = stime.split()
                if len(parts) == 2:
                    date, full_time = parts
                    time = full_time[:5]
                else:
                    date, time = 'Неизвестно', 'Неизвестно'
            else:
                date, time = 'Неизвестно', 'Неизвестно'
            program_data.append({
                'Сотрудник': fio,
                'Дата': date,
                'Время': time,
                'Программа': desc,
                'Сайт': ''
            })
    print(Fore.BLUE + 'load_program_data выполнено')
    return pd.DataFrame(program_data)


def load_site_data(output_file_path_site):
    """
    Загрузить данные о сайтах из XML файла.

    :param output_file_path_site: Путь к XML файлу, содержащему данные о пользователях и сайтах.
    :return: pandas DataFrame, содержащий информацию о сотрудниках, датах, времени, программах и сайтах
    """
    print(Fore.GREEN + 'Идет загрузка данных и сайтах из xml файла')
    """Загрузить данные о сайтах из XML файла."""
    parser = etree.XMLParser(remove_blank_text=True)
    tree = etree.parse(output_file_path_site, parser)
    root = tree.getroot()
    site_data = []
    for user in root.iter('user'):
        fio = user.find('fio')
        fio = fio.text if fio is not None else 'Неизвестно'
        for item in user.findall('item'):
            url = item.find('url')
            url = url.text if url is not None else 'Неизвестно'
            desc = item.find('desc')
            desc = desc.text if desc is not None else 'Неизвестно'
            stime = item.find('stime')
            stime = stime.text if stime is not None else 'Неизвестно'
            if stime != 'Неизвестно':
                parts = stime.split()
                if len(parts) == 2:
                    date, full_time = parts
                    time = full_time[:5]
                else:
                    date, time = 'Неизвестно', 'Неизвестно'
            else:
                date, time = 'Неизвестно', 'Неизвестно'
            site_data.append({
                'Сотрудник': fio,
                'Дата': date,
                'Время': time,
                'Программа': desc if desc != 'Неизвестно' else '',
                'Сайт': url
            })
    print(Fore.BLUE + 'load_site_data выполнено')
    return pd.DataFrame(site_data)


def update_employee_sheets(output_xlsx_path, program_data, site_data):
    """
    Обновить листы сотрудников в Excel файле данными о программах и сайтах.

    :param output_xlsx_path: Путь к исходному Excel файлу с листами сотрудников.
    :param program_data: DataFrame с данными о программах.
    :param site_data: DataFrame с данными о сайтах.
    """
    print(Fore.GREEN + 'Идет обновление excel фала с данными о программах и сайтах')
    workbook = load_workbook(output_xlsx_path)
    for sheet_name in workbook.sheetnames:
        employee_program_data = program_data[program_data['Сотрудник'] == sheet_name]
        employee_site_data = site_data[site_data['Сотрудник'] == sheet_name]
        if not employee_program_data.empty or not employee_site_data.empty:
            sheet = workbook[sheet_name]
            if sheet.max_row == 1:
                sheet.append(['Дата', 'Время', 'Программа', 'Сайт'])
            combined_data = pd.concat([employee_program_data, employee_site_data]).sort_values(by=['Дата', 'Время'])
            prev_date = None
            prev_time = None
            for _, row in combined_data.iterrows():
                current_date = row['Дата']
                current_time = row['Время']
                print_date = current_date if current_date != prev_date else ''
                print_time = current_time if current_date != prev_date or current_time != prev_time else ''
                sheet.append([print_date, print_time, row['Программа'], row['Сайт']])
                prev_date = current_date
                prev_time = current_time
    workbook.save(output_xlsx_path)
    print(f'Путь до обновленного и успешного сохраненного файла xlsx: {output_xlsx_path}')
    print(Fore.BLUE + 'update_employee_sheets и siteNprog_toexcel.py выполнены\n')
