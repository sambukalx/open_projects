"""
Все права защищены (c) 2024. 
Данный скрипт предназначен для работы с папками и файлами.
Автор кода не предоставляет прав на использование или распространение данного ПО.
"""

import os
import openpyxl
import sys
import traceback
import shutil
from colorama import init, Fore
from colorama import Style
import logging

logger = logging.getLogger(__name__)


def rename_folders_from_excel_cell(root_path):
    """
    :param root_path: Путь к корневому каталогу, где начинается процесс переименования папки.
    :return: Нет. Эта функция выполняет операции переименования непосредственно в файловой системе на основе содержимого файлов .xlsx.
    """
    print(Fore.GREEN + 'Идет переименование папок стахановца')
    if sys.stdin and hasattr(sys.stdin, 'reconfigure'):
        sys.stdin.reconfigure(encoding='utf-8')
    else:
        logger.info("sys.stdin недоступен или не поддерживает метод reconfigure.")
    if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    else:
        logger.info("sys.stdout недоступен или не поддерживает метод reconfigure.")
    for dirpath, dirnames, filenames in os.walk(root_path, topdown=False):
        xlsx_files = [f for f in filenames if f.lower().endswith('.xlsx')]
        if xlsx_files:
            first_xlsx_file = xlsx_files[0]
            xlsx_file_path = os.path.join(dirpath, first_xlsx_file)
            try:
                wb = openpyxl.load_workbook(xlsx_file_path, data_only=True)
                cell_value = None
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    cell_value_candidate = sheet['C2'].value
                    if cell_value_candidate:
                        cell_value = cell_value_candidate
                        break
                wb.close()
                if cell_value is None:
                    print(f"Ячейка C2 пуста на всех листах в файле '{xlsx_file_path}'. Пропускаем переименование.")
                    continue
                if not cell_value:
                    print(
                        f"Значение ячейки C2 после удаления недопустимых символов пустое. Пропускаем переименование '{dirpath}'.")
                    continue
                parent_dir = os.path.dirname(dirpath)
                new_folder_path = os.path.join(parent_dir, cell_value)
                if os.path.exists(new_folder_path):
                    continue
                os.rename(dirpath, new_folder_path)
            except Exception as e:
                print(f"Ошибка при обработке файла '{xlsx_file_path}': {e}")
                traceback.print_exc()
    print(Fore.BLUE + 'rename_folders_from_excel_cell и stahName.py выполнены\n')


# Функция для проверки и изменения значения в ячейке C2
def check_and_modify_excel(file_path, replacements):
    """
    :param file_path: Путь к файлу Excel, который необходимо проверить и изменить.
    :param replacements: Словарь, в котором ключи это значения, которые нужно заменить, а значения — это замены.
    :return: True, если значение ячейки было найдено и изменено или если ячейка уже содержит правильное значение, в противном случае — False.
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    cell_value = sheet['C2'].value
    if cell_value in replacements:
        sheet['C2'] = replacements[cell_value]
        wb.save(file_path)
        return True
    elif cell_value in replacements.values():
        return True
    else:
        return False


# Функция для удаления папки, если она содержит файлы, не удовлетворяющие условиям
def remove_folder_if_invalid(folder_path, invalid_xlsx_found):
    """
    :param folder_path: Путь к папке, которую необходимо проверить и, возможно, удалить.
    :param invalid_xlsx_found: Логический флаг, указывающий, были ли в папке обнаружены недействительные файлы xlsx.
    :return: None
    """
    if invalid_xlsx_found:
        try:
            shutil.rmtree(folder_path)
        except Exception as e:
            print(f'Ошибка при удалении папки {folder_path}: {e}')


# Основная функция для обхода папок и проверки файлов
def process_folders(base_path, replacements):
    """
    :param base_path: Путь к корневому каталогу, с которого начнется обход папки.
    :type base_path: str
    :param replacements: Словарь правил замены, которые будут применяться к файлам Excel.
    :type replacements: Dict
    :return: None
    :rtype: None
    """
    print(Fore.GREEN + 'Идет обход папок и проверка файлов')
    for root, dirs, files in os.walk(base_path, topdown=False):
        total_xlsx_files = 0
        invalid_xlsx_files = 0
        for file in files:
            if file.endswith('.xlsx'):
                total_xlsx_files += 1
                file_path = os.path.join(root, file)
                if not check_and_modify_excel(file_path, replacements):
                    invalid_xlsx_files += 1
        if total_xlsx_files > 0 and invalid_xlsx_files == total_xlsx_files:
            remove_folder_if_invalid(root, True)
    print(Fore.BLUE + 'process_folders в stahName.py выполнено\n')
