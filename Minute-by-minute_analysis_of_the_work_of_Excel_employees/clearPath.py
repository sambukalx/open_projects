"""
Все права защищены (c) 2024. 
Данный скрипт предназначен для очистки путей.
Автор кода не предоставляет прав на использование или распространение данного ПО.
"""

from colorama import init, Fore
from colorama import Style
import os
import shutil
import win32com.client as win32
from openpyxl import load_workbook
import xlrd
import pandas as pd


def delete_png_files(start_path):
    """
    :param start_path: корневой каталог, из которого следует начать поиск файлов для удаления.
    :return: None
    """
    print(Fore.GREEN + 'Идет удаление ненужных фалов из распакованного архива')
    for root, dirs, files in os.walk(start_path):
        for file in files:
            if file.lower().endswith('.png') or file.lower().endswith('.csv') or file.lower().endswith(
                    '.css') or file.lower().endswith('.html') or file.lower().endswith('.dat') or file.lower().endswith(
                    '.js') or file.lower().endswith('.htm'):
                file_path = os.path.join(root, file)
                os.remove(file_path)
    print(Fore.BLUE + 'delete_png_files выполнено')


def delete_small_folders(folder_path, min_files=5):
    """
    :param folder_path: Путь к папке для поиска и удаления мелких папок.
    :param min_files: Пороговый размер для удаления (в байтах).
    :return: None
    """
    print(Fore.GREEN + 'Идет удаление ненужных папок из распакованного архива')
    for root, dirs, files in os.walk(folder_path, topdown=False):
        for dir_name in dirs:
            dir_path = os.path.join(root, dir_name)
            try:
                num_files = sum([len(files) for _, _, files in os.walk(dir_path)])
                if num_files < min_files:
                    shutil.rmtree(dir_path)
            except Exception as e:
                print(f"Не удалось удалить папку {dir_path}: {e}")
    print(Fore.BLUE + 'delete_small_folders выполнено')


def convert_xls_to_xlsx(folder_path):
    """
    :param folder_path: Путь к директории, содержащей файлы XLS для конвертации.
    :return: None
    """
    print(Fore.GREEN + 'Идет конвертация файлов из распакованного архива')
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False  # Отключить предупреждения
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith('.xls'):
                xls_file = os.path.normpath(os.path.join(root, file))
                xlsx_file = os.path.splitext(xls_file)[0] + '.xlsx'
                try:
                    if os.path.exists(xls_file):
                        wb = excel.Workbooks.Open(xls_file)
                        wb.SaveAs(xlsx_file, FileFormat=51)  # 51 = xlOpenXMLWorkbook (формат .xlsx)
                        wb.Close()
                    else:
                        print(f"Файл не найден: {xls_file}")
                except Exception as e:
                    print(f"Ошибка при конвертации файла {xls_file}: {e}")
    excel.Quit()
    print(Fore.BLUE + 'convert_xls_to_xlsx выполнено')


def delete_x_files(folder_path):
    """
    :param folder_path: Путь к папке, в которой нужно удалить файлы с расширением xls.
    :return: None
    """
    print(Fore.GREEN + 'Идет удаление ненужных фалов из распакованного архива')
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.lower().endswith('.xls'):
                file_path = os.path.join(root, file)
                os.remove(file_path)
    print(Fore.BLUE + 'delete_x_files выполнено')


def delete_folders_based_on_C2_recursive(start_path, word='телефон'):
    """
    :param start_path: Путь к корневому каталогу, с которого начнется процесс поиска и удаления.
    :param word: Слово для поиска в ячейке C2 любого найденного файла Excel. Если ячейка пуста или содержит это слово, папка будет удалена.
    :return: None
    """
    print(Fore.GREEN + 'Идет удаление ненужных папок из распакованного архива')
    for root, dirs, files in os.walk(start_path, topdown=False):
        excel_files = [f for f in files if f.lower().endswith(('.xlsx', '.xls'))]
        if excel_files:
            first_excel = os.path.join(root, excel_files[0])
            try:
                cell_value = None
                if first_excel.lower().endswith('.xlsx'):
                    wb = load_workbook(first_excel, read_only=True, data_only=True)
                    ws = wb.active
                    cell_value = ws['C2'].value
                    wb.close()
                elif first_excel.lower().endswith('.xls'):
                    workbook = xlrd.open_workbook(first_excel)
                    sheet = workbook.sheet_by_index(0)
                    cell_value = sheet.cell_value(1, 2)
                if (cell_value is None) or (str(cell_value).strip() == '') or (word.lower() in str(cell_value).lower()):
                    shutil.rmtree(root)
                    dirs[:] = []
            except Exception as e:
                print(f"Ошибка при обработке файла {first_excel}: {e}")
        else:
            pass
    print(Fore.BLUE + 'delete_folders_based_on_C2_recursive выполнено')


def delete_pacs(start_path):
    """
    :param start_path: Путь к каталогу, из которого следует начать поиск файлов для удаления.
    :return: None
    """
    print(Fore.GREEN + 'Идет удаление ненужных фалов из распакованного архива')
    for root, dirs, files in os.walk(start_path):
        for file in files:
            if 'pacs' in file and file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                os.remove(file_path)
    print(Fore.BLUE + 'delete_pacs и clearPath.py выполнены\n')
