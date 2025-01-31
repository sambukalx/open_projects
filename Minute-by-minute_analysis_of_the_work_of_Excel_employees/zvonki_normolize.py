"""
Все права защищены (c) 2024. 
Данный скрипт предназначен для нормализации информации.
Автор кода не предоставляет прав на использование или распространение данного ПО.
"""

import shutil
import os
from colorama import init, Fore
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


def create_file_copy(original_path):
    """
    :param original_path: Путь к исходному файлу, который необходимо скопировать.
    :return: Путь к созданной копии файла.
    """
    print(Fore.GREEN + 'Идет создание копии файла со звонками')
    global copy_file_path_zv
    directory, original_filename = os.path.split(original_path)
    filename_without_ext, ext = os.path.splitext(original_filename)
    copy_filename = f"{filename_without_ext}_копия{ext}"
    copy_file_path_zv = os.path.join(directory, copy_filename)
    try:
        shutil.copyfile(original_path, copy_file_path_zv)
        copy_file_path_zv = copy_file_path_zv.replace('\\', '/')
    except Exception as e:
        print(f"Ошибка при создании копии файла: {e}")
        raise e
    print(f'Путь до копии фала АТС: {copy_file_path_zv}')
    print(Fore.BLUE + 'create_file_copy выполнено')
    return copy_file_path_zv


def process_and_save_calls_data(file_path, replacements):
    """
    :param file_path: Путь к обрабатываемому Excel файлу.
    :param replacements: Словарь замен, применяемых к указанному столбцу.
    :return: None
    """
    print(Fore.GREEN + 'Идет переоформление данных внутри копии файла со звонками')
    df = pd.read_excel(file_path, skiprows=8, header=1)
    df = df[df['Тип звонка'] != 'пропущенный']
    df.iloc[:, 2] = df.iloc[:, 2].replace(replacements)
    column_names = df.columns.tolist()
    col_J_name = column_names[9]
    col_K_name = column_names[10]
    df[col_J_name] = pd.to_timedelta(df[col_J_name].astype(str))
    df[col_K_name] = pd.to_timedelta(df[col_K_name].astype(str))
    df['Combined_Time'] = df[col_J_name] + df[col_K_name]
    df['Combined_Time'] = df['Combined_Time'].apply(
        lambda x: f"{int(x.total_seconds() // 60):02}:{int(x.total_seconds() % 60):02}")
    columns_to_drop_indices = [3, 5, 6, 11]
    columns_to_drop_names = [df.columns[i] for i in columns_to_drop_indices if i < len(df.columns)]
    columns_to_drop_names.extend([col_J_name, col_K_name])
    df = df.drop(columns=columns_to_drop_names)
    df = df.rename(columns={'Combined_Time': 'Длительность'})
    df.to_excel(file_path, index=False)
    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
        for cell in row:
            if cell.column_letter in ['B', 'D']:
                cell.number_format = '00000000000'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.value = cell.value.strftime('%Y-%m-%d')
            elif isinstance(cell.value, str) and ' ' in cell.value:
                try:
                    date_value = datetime.strptime(cell.value.split(' ')[0], '%d.%m.%Y')
                    cell.value = date_value.strftime('%Y-%m-%d')
                except Exception as e:
                    print(f"Ошибка при преобразовании значения {cell.value}: {e}")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.value = cell.value.strftime('%H:%M')
            elif isinstance(cell.value, str) and ':' in cell.value:
                try:
                    time_value = datetime.strptime(cell.value, '%H:%M:%S')
                    cell.value = time_value.strftime('%H:%M')
                except Exception as e:
                    print(f"Ошибка при преобразовании значения {cell.value}: {e}")
    wb.save(file_path)
    print(f"Файл успешно сохранен: {file_path}")
    print(Fore.BLUE + 'process_and_save_calls_data и zvonki_normolize.py выполнены\n')
