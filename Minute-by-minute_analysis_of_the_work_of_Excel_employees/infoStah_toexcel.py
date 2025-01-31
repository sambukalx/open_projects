"""
Все права защищены (c) 2024. 
Данный скрипт предназначен для обработки Excel файла.
Автор кода не предоставляет прав на использование или распространение данного ПО.
"""

import openpyxl
from datetime import datetime, timedelta
from colorama import init, Fore
import re


def format_time_string(time_str):
    """
    :param time_str: строка, представляющая продолжительность, содержащую часы и/или минуты.
    :return: Отформатированная строка, представляющая продолжительность в формате «ЧЧ:ММ» или «П», если синтаксический анализ не удался.
    """
    hours = 0
    minutes = 0
    hours_match = re.search(r'(\d+)\s*ч', time_str)
    minutes_match = re.search(r'(\d+)\s*м', time_str)
    if hours_match:
        try:
            hours = int(hours_match.group(1))
        except ValueError:
            hours = 0
    if minutes_match:
        try:
            minutes = int(minutes_match.group(1))
        except ValueError:
            minutes = 0
    if hours == 0 and minutes == 0 and time_str.strip() != "":
        return "П"
    return f"{hours:02d}:{minutes:02d}"


def update_excel_with_employee_data(info_work_stah, output_file_excelSite):
    """
    :param info_work_stah: Словарь, где ключами являются имена сотрудников, а значениями являются словари, содержащие их рабочие данные, включая pc_time, время входа и time_act.
    :param output_file_excelSite: Путь к файлу Excel, в котором будут обновляться данные о сотрудниках.
    :return: None
    """
    workbook = openpyxl.load_workbook(output_file_excelSite)
    for employee, data in info_work_stah.items():
        if employee in workbook.sheetnames:
            sheet = workbook[employee]
            for month, days_data in data['pc_time'].items():
                for day, day_info in days_data.items():
                    date_str = f"2024-{month.zfill(2)}-{str(day).zfill(2)}"
                    try:
                        date = datetime.strptime(date_str, '%Y-%m-%d')
                    except ValueError:
                        continue
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                        cell_value = row[0].value
                        if isinstance(cell_value, str):
                            try:
                                cell_date = datetime.strptime(cell_value, '%Y-%m-%d').date()
                            except ValueError:
                                cell_date = None
                        elif isinstance(cell_value, datetime):
                            cell_date = cell_value.date()
                        else:
                            cell_date = cell_value
                        # Сравниваем даты
                        if cell_date == date.date():
                            row_index = row[0].row
                            # Запись данных о времени входа/выхода
                            enter_time = data['enterexit'].get(month, {}).get(day, {}).get('start', {}).get('means', '')
                            exit_time = data['enterexit'].get(month, {}).get(day, {}).get('end', {}).get('meane', '')
                            if enter_time and exit_time:
                                sheet.cell(row=row_index, column=9).value = f"{enter_time} - {exit_time}"
                            elif enter_time:
                                sheet.cell(row=row_index, column=9).value = f"{enter_time}"
                            elif exit_time:
                                sheet.cell(row=row_index, column=9).value = f"{exit_time}"
                            # Время работы компьютера и активное время
                            pc_time = day_info.get('mean', '')
                            if pc_time:
                                sheet.cell(row=row_index, column=10).value = format_time_string(pc_time)
                            active_time = data['time_act'].get(month, {}).get(day, {}).get('meant', '')
                            if active_time:
                                sheet.cell(row=row_index, column=6).value = format_time_string(active_time)
    workbook.save(output_file_excelSite)
    print(f"Файл сохранён: {output_file_excelSite}")
