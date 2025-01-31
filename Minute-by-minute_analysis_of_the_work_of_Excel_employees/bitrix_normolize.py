"""
Все права защищены (c) 2024. 
Данный скрипт предназначен для сбора информации из Битрикс24.
Автор кода не предоставляет прав на использование или распространение данного ПО.
"""

import pandas as pd
import openpyxl
from datetime import datetime, timedelta


def convert_html_to_xlsx(file_path):
    """
    :param file_path: Путь к HTML-файлу, который необходимо преобразовать.
    :return: Путь к вновь созданному файлу XLSX.
    """
    tables = pd.read_html(file_path)
    file_path_bit = file_path.replace('.xls', '.xlsx')
    tables[0].to_excel(file_path_bit, index=False)
    print(f"Файл успешно конвертирован из HTML в XLSX: {file_path_bit}")
    return file_path_bit


def format_time_data(time_string):
    """
    :param time_string: строка, представляющая время в формате «часы минуты начальное_время — конечное_время». Пример: «7 26 10:33 – 18:00»
    :return: Отформатированная строка в формате «часы:минуты время_начала — время_окончания». Пример: «7:26 10:33 — 18:00», если входная строка
             находится в ожидаемом формате; в противном случае возвращает исходную входную строку.
    """
    parts = time_string.split()
    if len(parts) == 5:
        hours = parts[0]
        minutes = parts[1]
        start_time = parts[2]
        end_time = parts[4]
        return f"{hours}:{minutes} {start_time} - {end_time}"
    return time_string


def adjust_time_for_timezone(start_time, end_time, timezone_difference):
    """
    :param start_time: Время начала в виде строки в формате «ЧЧ:ММ».
    :param end_time: Время окончания в виде строки в формате «ЧЧ:ММ».
    :param timezone_difference: Часовой пояс компании.
    :return: Start_time и end_time в строковом формате с часовым поясом компании.
    """
    start_time = datetime.strptime(start_time, '%H:%M') + timedelta(hours=timezone_difference)
    end_time = datetime.strptime(end_time, '%H:%M') + timedelta(hours=timezone_difference)
    return start_time.strftime('%H:%M'), end_time.strftime('%H:%M')


def replace_values_in_xlsx(file_path_bit, replacements):
    """
    :param file_path_bit:  Путь к файлу .xlsx, который будет обработан.
    :type file_path_bit: str
    :param replacements: Словарь значений, которые нужно заменить в файле .xlsx.
    :type replacements: dict
    :return: None
    :rtype: None
    """
    workbook = openpyxl.load_workbook(file_path_bit)
    sheet = workbook.active
    for row in range(1, 101):
        employee_name = None
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell_value = str(cell.value)
            if cell_value in employee_timezones:
                employee_name = cell_value
            if cell_value in replacements:
                cell.value = replacements[cell_value]
            if "-" in cell_value:
                formatted_time = format_time_data(cell_value)
                cell.value = formatted_time
                if employee_name:
                    timezone_difference = company_timezone - employee_timezones[employee_name]
                    start_time_str, end_time_str = formatted_time.split(' ')[1:3]
                    start_time_adjusted, end_time_adjusted = adjust_time_for_timezone(start_time_str, end_time_str, timezone_difference)
                    cell.value = f"{formatted_time.split()[0]} {start_time_adjusted} - {end_time_adjusted}"
    workbook.save(file_path_bit)
    print(f"Значения успешно заменены и отформатированы в файле: {file_path_bit}")
