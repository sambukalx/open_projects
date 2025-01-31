"""
Все права защищены (c) 2024. 
Данный скрипт предназначен для работы с Excel файлами.
Автор кода не предоставляет прав на использование или распространение данного ПО.
"""

from colorama import init, Fore
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, time, date, timedelta


def zvonkiExcel(output_file_excelSite, copy_file_path_zv):
    """
    :param output_file_excelSite: Путь к выходному Excel файлу, куда будут сохранены преобразованные данные звонков.
    :param copy_file_path_zv: Путь к оригинальному Excel файлу, содержащему исходные данные звонков.
    :return: Преобразованный и обновленный Excel файл с данными звонков, отсортированными и вставленными в соответствующие листы для каждого сотрудника.
    """
    print(Fore.GREEN + 'Идет запись данных о звонках в excel файл')
    original_wb = load_workbook(copy_file_path_zv)
    reports_wb = load_workbook(output_file_excelSite)
    original_sheet = original_wb['Sheet1']
    call_data = []
    for row in original_sheet.iter_rows(min_row=2, values_only=True):
        row_data = list(row)
        if len(row_data) >= 7:
            call_type = row_data[0]
            client = row_data[1]
            employee = row_data[2]
            through = row_data[3]
            date_val = row_data[4]
            time_val = row_data[5]
            duration_val = row_data[6]
            try:
                parsed_date = pd.to_datetime(date_val).strftime('%Y-%m-%d') if date_val else None
            except (ValueError, TypeError):
                parsed_date = None
            try:
                parsed_time = pd.to_datetime(time_val).strftime('%H:%M') if time_val else None
            except (ValueError, TypeError):
                parsed_time = None
            try:
                if isinstance(duration_val, str):
                    parts = duration_val.split(':')
                    if len(parts) == 2:
                        minutes, seconds = map(int, parts)
                        parsed_duration = timedelta(minutes=minutes, seconds=seconds)
                    elif len(parts) == 3:
                        hours, minutes, seconds = map(int, parts)
                        parsed_duration = timedelta(hours=hours, minutes=minutes, seconds=seconds)
                    else:
                        parsed_duration = timedelta(0)
                elif isinstance(duration_val, (int, float)):
                    parsed_duration = timedelta(seconds=duration_val)
                else:
                    parsed_duration = timedelta(0)
                total_minutes = int(parsed_duration.total_seconds() // 60)
                if parsed_duration.total_seconds() % 60 > 0:
                    total_minutes += 1
            except (ValueError, TypeError) as e:
                print(f"Ошибка при обработке длительности звонка: {duration_val}, ошибка: {e}")
                parsed_duration = timedelta(0)
                total_minutes = 1
            call_data.append({
                'Тип звонка': call_type,
                'Клиент': client,
                'Сотрудник': employee,
                'Через': through,
                'Дата': parsed_date,
                'Время': parsed_time,
                'Длительность': parsed_duration,
                'Минуты': total_minutes
            })
    employee_data = {sheet: [] for sheet in reports_wb.sheetnames if
                     sheet not in ['Отдел продаж (ОП)', 'Отдел продаж 2 (ОП2)']}
    for call in call_data:
        employee_name = call['Сотрудник']
        if employee_name in employee_data:
            employee_data[employee_name].append(call)
    print(Fore.GREEN + 'Идет преобразование данных на листах сотрудников в подходящий формат')

    def convert_sheet_data(sheet):
        """
        :param sheet: Excel-лист, из которого будут извлечены данные.
        :return: Словарь, отображающий пары дата-время на строки в листе.
        """
        date_time_map = {}
        current_date = None
        for row in range(2, sheet.max_row + 1):
            sheet_date_cell = sheet.cell(row=row, column=1)
            sheet_time_cell = sheet.cell(row=row, column=2)
            if sheet_date_cell.value:
                if isinstance(sheet_date_cell.value, (datetime, date)):
                    current_date = sheet_date_cell.value.strftime('%Y-%m-%d')
                elif isinstance(sheet_date_cell.value, str):
                    current_date = sheet_date_cell.value.strip()
                else:
                    current_date = None
            sheet_time = None
            if sheet_time_cell.value:
                if isinstance(sheet_time_cell.value, (datetime, time)):
                    sheet_time = sheet_time_cell.value.strftime('%H:%M')
                elif isinstance(sheet_time_cell.value, str):
                    sheet_time = sheet_time_cell.value.strip()
                else:
                    sheet_time = None
            if current_date and sheet_time:
                if (current_date, sheet_time) not in date_time_map:
                    date_time_map[(current_date, sheet_time)] = []
                date_time_map[(current_date, sheet_time)].append(row)
        return date_time_map
    print(Fore.BLUE + 'convert_sheet_data выполнено')
    print(Fore.GREEN + 'Идет вставка данных о звонках в соответствующие листы сотрудников')

    def insert_call_data_to_report(employee_data, reports_wb):
        """
        Вставляет данные о звонках в рабочую книгу отчета для каждого сотрудника.

        :param employee_data: Словарь, где ключ — имя сотрудника, а значение — список звонков.
                              Каждый звонок представлен как словарь с ключами 'Дата', 'Время', 'Минуты' и т.д.
        :param reports_wb: Объект рабочей книги, содержащий листы для каждого сотрудника.
        :return: None
        """
        for employee_name, calls in employee_data.items():
            employee_sheet = reports_wb[employee_name]
            if employee_sheet.cell(row=1, column=5).value != 'Звонки':
                employee_sheet.cell(row=1, column=5, value='Звонки')
            date_time_map = convert_sheet_data(employee_sheet)
            calls.sort(key=lambda x: (x['Дата'], x['Время']))
            for call in calls:
                call_date = call['Дата']
                call_time = call['Время']
                minutes = call.get('Минуты', 1)
                start_datetime = pd.to_datetime(f"{call_date} {call_time}")
                call_times = [start_datetime + timedelta(minutes=i) for i in range(minutes)]
                for idx, call_time_dt in enumerate(call_times):
                    call_date_str = call_time_dt.strftime('%Y-%m-%d')
                    call_time_str = call_time_dt.strftime('%H:%M')
                    key = (call_date_str, call_time_str)
                    call_entry = call.copy()
                    call_entry['Дата'] = call_date_str
                    call_entry['Время'] = call_time_str
                    if idx == 0:
                        call_entry['Тип звонка'] = call['Тип звонка']
                    else:
                        call_entry['Тип звонка'] = f"Продолжение звонка: {call['Тип звонка']}"
                    if key in date_time_map:
                        last_row = date_time_map[key][-1]
                        insert_new_row_below(employee_sheet, last_row, call_entry, date_time_map)
                    else:
                        insert_new_row_sorted(employee_sheet, call_entry, date_time_map)
    print(Fore.BLUE + 'insert_call_data_to_report выполнено')
    print(Fore.GREEN + 'Идет запись данных о звонках в excel файл')

    def insert_new_row_below(sheet, row, call, date_time_map):
        """
        Вставляет новую строку ниже указанной и обновляет date_time_map.

        :param sheet: Excel-лист, в который будет вставлена новая строка.
        :param row: Индекс строки, ниже которой будет вставлена новая строка.
        :param call: Словарь с информацией о звонке.
        :param date_time_map: Словарь, отображающий пары дата-время на списки индексов строк.
                              Будет обновлен для отражения нового индекса строки.
        :return: None
        """
        new_row = row + 1
        sheet.insert_rows(new_row)
        for key in date_time_map:
            updated_rows = []
            for r in date_time_map[key]:
                if r >= new_row:
                    updated_rows.append(r + 1)
                else:
                    updated_rows.append(r)
            date_time_map[key] = updated_rows
        key = (call['Дата'], call['Время'])
        if key not in date_time_map:
            date_time_map[key] = []
        date_time_map[key].append(new_row)
        date_time_map[key].sort()
        sheet.cell(row=new_row, column=1, value=call['Дата'])
        sheet.cell(row=new_row, column=2, value=call['Время'])
        sheet.cell(row=new_row, column=5, value=f"{call['Тип звонка']} к {call['Клиент']}")
    print(Fore.BLUE + 'insert_new_row_below выполнено')
    print(Fore.GREEN + 'Идет запись данных о звонках в excel файл последнее')

    def insert_new_row_sorted(sheet, call, date_time_map):
        """
        Вставляет новую строку в отсортированном порядке и обновляет date_time_map.

        :param sheet: Excel-лист, в который будет вставлена новая строка.
        :param call: Словарь с информацией о звонке.
        :param date_time_map: Словарь, отображающий пары дата-время на списки индексов строк.
                              Будет обновлен для отражения нового индекса строки.
        :return: None
        """
        call_date = call['Дата']
        call_time = call['Время']
        all_date_times = list(date_time_map.keys())
        all_date_times.append((call_date, call_time))
        all_date_times = sorted(all_date_times)
        position = all_date_times.index((call_date, call_time))
        if position == 0:
            new_row = 2
        else:
            prev_key = all_date_times[position - 1]
            prev_rows = date_time_map[prev_key]
            new_row = prev_rows[-1] + 1
        sheet.insert_rows(new_row)
        for key in date_time_map:
            updated_rows = []
            for r in date_time_map[key]:
                if r >= new_row:
                    updated_rows.append(r + 1)
                else:
                    updated_rows.append(r)
            date_time_map[key] = updated_rows
        if (call_date, call_time) not in date_time_map:
            date_time_map[(call_date, call_time)] = []
        date_time_map[(call_date, call_time)].append(new_row)
        date_time_map[(call_date, call_time)].sort()
        sheet.cell(row=new_row, column=1, value=call_date)
        sheet.cell(row=new_row, column=2, value=call_time)
        sheet.cell(row=new_row, column=5, value=f"{call['Тип звонка']} к {call['Клиент']}")
    print(Fore.BLUE + 'insert_new_row_sorted выполнено')
    print(Fore.GREEN + 'Идет вставка данных о звонках в соответствующие листы сотрудников')
    insert_call_data_to_report(employee_data, reports_wb)
    reports_wb.save(output_file_excelSite)
    print(f'Путь до обработанного файла с программами, сайтами и звонками: {output_file_excelSite}')
    print(Fore.BLUE + 'zvonkiExcel.py выполнено\n')
