"""
Все права защищены (c) 2024. 
Данный скрипт предназначен для форматирования Excel файла.
Автор кода не предоставляет прав на использование или распространение данного ПО.
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from colorama import init, Fore


def format_excel_file(file_path):
    """
    :param file_path: Путь к файлу Excel, который необходимо отформатировать.
    :return: None
    """
    print(Fore.GREEN + 'Идет форматирование файла отчета')
    workbook = openpyxl.load_workbook(file_path)

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    top_border = Border(top=Side(style='thin'))
    bottom_border = Border(bottom=Side(style='thin'))
    all_borders = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    updown_border = Border(top=Side(style='thin'),
                           bottom=Side(style='thin'))
    incoming_call_font = Font(color="00B0F0")
    outgoing_call_font = Font(color="00B050")
    unsuccessful_call_font = Font(color="FF0000")
    telegram_font = Font(color="00B0F0")
    whatsapp_font = Font(color="008000")
    youtube_font = Font(color="FF0000")
    headers = [
        'Дата', 'Время', 'Программа', 'Сайт', 'Звонки',
        'Активное время Стахановец', 'Активное время Битрикс',
        'Вход/выход Битрикс', 'Вход/выход Стахановец', 'Время работы ПК'
    ]
    column_widths = [10.5, 10.5, 33, 26, 37, 32, 32, 25, 25, 18]
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith('Отдел') or sheet_name.startswith('отдел'):
            print(f"Лист '{sheet_name}' пропущен.")
            continue
        sheet = workbook[sheet_name]
        sheet.freeze_panes = sheet['A2']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = thin_border
        for col_num, width in enumerate(column_widths, 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = width
        seen_dates = set()
        for row in range(2, sheet.max_row + 1):
            date_cell = sheet.cell(row=row, column=1)
            if date_cell.value in seen_dates:
                date_cell.value = None
            else:
                seen_dates.add(date_cell.value)
                date_cell.font = bold_font
        current_date = None
        start_row = None
        # Пройдем по всем строкам с данными
        for row in range(2, sheet.max_row + 1):
            date_cell = sheet.cell(row=row, column=1)
            if date_cell.value is not None:
                if current_date is not None and start_row is not None and start_row < row - 1:
                    sheet.row_dimensions.group(start_row, row - 1, hidden=True)
                current_date = date_cell.value
                start_row = row + 1
            elif start_row is None:
                start_row = row
        if current_date is not None and start_row is not None and start_row <= sheet.max_row:
            if start_row <= sheet.max_row:
                sheet.row_dimensions.group(start_row, sheet.max_row, hidden=True)
        seen_fij = set()
        for row in range(2, sheet.max_row + 1):
            f_value = sheet.cell(row=row, column=6).value
            i_value = sheet.cell(row=row, column=9).value
            j_value = sheet.cell(row=row, column=10).value
            fij_combination = (f_value, i_value, j_value)
            if fij_combination in seen_fij:
                sheet.cell(row=row, column=6).value = None
                sheet.cell(row=row, column=9).value = None
                sheet.cell(row=row, column=10).value = None
            else:
                seen_fij.add(fij_combination)
        for row in range(2, sheet.max_row + 1):
            a_cell = sheet.cell(row=row, column=1)
            if a_cell.value:
                a_cell.border = top_border
            b_cell = sheet.cell(row=row, column=2)
            if b_cell.value:
                b_cell.border = top_border
            c_cell = sheet.cell(row=row, column=3)
            d_cell = sheet.cell(row=row, column=4)
            c_cell.border = bottom_border
            d_cell.border = bottom_border
            e_cell = sheet.cell(row=row, column=5)
            if e_cell.value:
                e_cell.border = updown_border
            for col in [6, 7, 8, 9, 10]:
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    cell.border = all_borders
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                if cell.column == 5 and cell.value:
                    if cell.value.startswith("входящий"):
                        cell.font = incoming_call_font
                    elif cell.value.startswith("исходящий"):
                        cell.font = outgoing_call_font
                    elif cell.value.startswith("неуспешный"):
                        cell.font = unsuccessful_call_font
                if cell.column == 4 and cell.value:
                    if "Telegramm" in cell.value:
                        cell.font = telegram_font
                    elif "WhatsApp" in cell.value:
                        cell.font = whatsapp_font
                    elif "YouTube" in cell.value:
                        cell.font = youtube_font
        sheet.auto_filter.ref = f"A1:J{sheet.max_row}"
        sheet.delete_cols(7, 2)
    workbook.save(file_path)
    print(f"Файл '{file_path}' успешно отформатирован.")
    print(Fore.BLUE + 'format_excel_file и format.py выполнены\n')
